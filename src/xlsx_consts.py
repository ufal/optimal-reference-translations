from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Color, PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from collections import defaultdict

FILL_0A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('aaaaaa')
)
FILL_0B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('dddddd')
)
FILL_1A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('ffc488')
)
FILL_1B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('fceee0')
)
FILL_2A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('8574fb')
)
FILL_2B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('eae8fc')
)
FILL_3A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('e9ff6d')
)
FILL_3B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('fcfcdb')
)
FILL_4A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('3c974b')
)
FILL_4B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('cbf2d1')
)
FILL_BLANK = PatternFill(
    patternType="solid",
    fill_type='solid', fgColor=Color('eda1a1')
)


THIN_BORDER_ALL = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
MEDIUM_BORDER_RIGHT = Border(
    left=Side(style='thin'),
    right=Side(style='medium'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
MEDIUM_BORDER_ALL = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)
THICK_BORDER_RIGHT = Border(
    left=Side(style='thin'),
    right=Side(style='thick'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
THICK_BORDER_BOTTOM = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thick')
)

VALIDATION_NUM = defaultdict(lambda: DataValidation(
    type="decimal",
    operator="between",
    formula1=0,
    formula2=6,
    errorStyle="stop",
))

VALIDATION_NONE = defaultdict(lambda: DataValidation(
    type="textLength",
    operator="between",
    formula1=0,
    formula2=0,
    errorStyle="stop",
))

FONT_BOLD = Font(bold=True, name="Arial")
FONT_NORMAL = Font(bold=False, name="Arial")

FORMATTING_BLANK = lambda cell: FormulaRule(formula=[f'ISBLANK({cell})'], stopIfTrue=True, fill=FILL_BLANK)


def ord_to_col(i):
    if i < 0:
        raise Exception(f"Attempted to transform {i} into a column name")
    if i < 26:
        return chr(ord("A") + i)
    if i < 26 * 26:
        return ord_to_col(i // 26 - 1) + ord_to_col(i % 26)
    raise Exception(f"Too large a column number {i}")


def get_height_for_row(sheet, row_number):
    row = list(sheet.rows)[row_number - 1]
    return max(
        40,
        max([
            len(cell.value) / 2.7 if cell.value is not None else 0
            for cell in row
        ])
    )
