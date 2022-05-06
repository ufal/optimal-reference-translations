#!/usr/bin/env python3

from utils import load_data, load_data_structure, save_json
import csv
from argparse import ArgumentParser
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, PatternFill
import random
from collections import defaultdict
from xlsx_consts import *

args = ArgumentParser()
args = args.parse_args()

ATTRIBUTES = [
    "Spelling", "Terminology", "Grammar",
    "Meaning", "Style", "Pragmatics", "Overall"
]


def ord_to_col(i):
    if i < 0:
        raise Exception(f"Attempted to transform {i} into a column name")
    if i < 26:
        return chr(ord("A") + i)
    if i < 26 * 26:
        return ord_to_col(i // 26 - 1) + ord_to_col(i % 26)
    raise Exception(f"Too large a column number {i}")


def get_height_for_row(sheet, row_number):
    row = list(sheet.rows)[row_number - 1]
    return max(
        40,
        max([
            len(cell.value) / 2.7 if cell.value is not None else 0
            for cell in row
        ])
    )


COLS_ATTRIBUTES_ALL = [
    [
        ord_to_col(1 + j + (len(ATTRIBUTES) + 1) * i)
        for j in range(len(ATTRIBUTES) + 1)
    ]
    for i in range(4)
]
COLS_ATTRIBUTES = [
    [
        ord_to_col(2 + j + (len(ATTRIBUTES) + 1) * i)
        for j in range(len(ATTRIBUTES))
    ]
    for i in range(4)
]
COLS_ATTRIBUTES_FLAT = {
    x
    for y in COLS_ATTRIBUTES
    for x in y
}

COLS_TRANSLATIONS = {"A"} | {
    ord_to_col(1 + i * (len(ATTRIBUTES) + 1))
    for i in range(4)
}
COLS_TRANSLATIONS_TRUE = [
    ord_to_col(1 + i * (len(ATTRIBUTES) + 1))
    for i in range(4)
]

COL_LAST = ord_to_col(1 + (len(ATTRIBUTES) + 1) * 4)
COLS_ALL = COLS_ATTRIBUTES_FLAT | COLS_TRANSLATIONS | {COL_LAST}


def add_edit_sheet(workbook, doc_i, doc_k, doc_v):
    sheet = workbook.create_sheet("Edit" + str(doc_i))
    sheet.add_data_validation(VALIDATION_NUM[doc_k])
    sheet.add_data_validation(VALIDATION_NONE[doc_k])

    for col in COLS_ALL:
        sheet[col + "1"].font = FONT_BOLD

    # header styling
    sheet["A1"].value = "Source"
    sheet["A1"].fill = FILL_0A
    for i, col in enumerate(COLS_TRANSLATIONS_TRUE):
        i = i + 1
        sheet[f"{col}1"].value = f"Translation {i}"

    for cols, style in zip(COLS_ATTRIBUTES_ALL, [FILL_1A, FILL_2A, FILL_3A, FILL_4A]):
        for col in cols:
            sheet[col + "1"].fill = style

    for cols_i, cols in enumerate(COLS_ATTRIBUTES):
        for col, attribute in zip(cols, ATTRIBUTES):
            sheet[f"{col}1"].value = f"T{cols_i} {attribute}"

    sheet[f"{COL_LAST}1"].value = "Comments"
    sheet.column_dimensions[COL_LAST].width = 30

    for col in COLS_ATTRIBUTES_FLAT:
        sheet[col + "1"].alignment = Alignment(
            textRotation=90, horizontal="center"
        )
        sheet.column_dimensions[col].width = 3
    for col in COLS_ALL:
        sheet[col + "1"].border = THICK_BORDER_BOTTOM

    sheet.row_dimensions[1].height = 85
    sheet.freeze_panes = sheet["B2"]

    # fill values
    for line_i, line in enumerate(doc_v):
        line_i += 2
        for sent_i, (sent, col) in enumerate(zip(line, ["A"] + COLS_TRANSLATIONS_TRUE)):
            cell = sheet[col + str(line_i)]
            if col == "A" or (line_i - 1 >= DOC_SPANS[doc_k][0] and line_i - 1 <= DOC_SPANS[doc_k][1]):
                cell.value = sent
            
            for col in COLS_ATTRIBUTES_FLAT:
                if line_i - 1 >= DOC_SPANS[doc_k][0] and line_i - 1 <= DOC_SPANS[doc_k][1]:
                    sheet.conditional_formatting.add(col + str(line_i), FORMATTING_BLANK(col + str(line_i)))
                    VALIDATION_NUM[doc_k].add(col + str(line_i))
                else:
                    VALIDATION_NONE[doc_k].add(col + str(line_i))

            cell.alignment = Alignment(wrap_text=True)

        if line_i % 2 == 0:
            sheet["A" + str(line_i)].fill = FILL_0B
            for col in COLS_ATTRIBUTES_ALL[0]:
                sheet[col + str(line_i)].fill = FILL_1B
            for col in COLS_ATTRIBUTES_ALL[1]:
                sheet[col + str(line_i)].fill = FILL_2B
            for col in COLS_ATTRIBUTES_ALL[2]:
                sheet[col + str(line_i)].fill = FILL_3B
            for col in COLS_ATTRIBUTES_ALL[3]:
                sheet[col + str(line_i)].fill = FILL_4B

        # set borders & alignment
        for col in COLS_ATTRIBUTES_FLAT | COLS_TRANSLATIONS:
            sheet[col + str(line_i)].border = THIN_BORDER_ALL
            sheet[col + str(line_i)].font = FONT_NORMAL
        for col in COLS_TRANSLATIONS:
            sheet[col + str(line_i)].alignment = Alignment(
                vertical="top", horizontal="justify", wrapText=True,
            )
        for cols in COLS_ATTRIBUTES_ALL:
            sheet[cols[-1] + str(line_i)].border = MEDIUM_BORDER_RIGHT

        # add data validation
        for col in COLS_ATTRIBUTES_FLAT:
            sheet[col + str(line_i)].alignment = Alignment(
                textRotation=90, horizontal="center"
            )

        sheet["A" + str(line_i)].border = THICK_BORDER_RIGHT

        sheet.row_dimensions[line_i].height = get_height_for_row(sheet, line_i)

    line_i += 2
    sheet["A" + str(line_i)].value = f"Document rating"
    sheet["A" + str(line_i+1)].value = f"Document time (minutes)"
    sheet.conditional_formatting.add("B" + str(line_i+1), FORMATTING_BLANK(col + str(line_i+1)))
    sheet["B" + str(line_i+1)].border = MEDIUM_BORDER_ALL
    sheet["A" + str(line_i+1)].fill = FILL_0A
    sheet["A" + str(line_i+1)].font = FONT_BOLD
    sheet.row_dimensions[line_i].height = 40

    for col in COLS_TRANSLATIONS - {"A"}:
        VALIDATION_NONE[doc_k].add(col + str(line_i))

    sheet["A" + str(line_i)].border = MEDIUM_BORDER_ALL
    sheet["A" + str(line_i)].font = FONT_BOLD
    sheet["A" + str(line_i)].fill = FILL_0A

    for col in COLS_ATTRIBUTES_FLAT:
        sheet[col + str(line_i)].border = MEDIUM_BORDER_ALL
        VALIDATION_NUM[doc_k].add(col + str(line_i))
        sheet.conditional_formatting.add(col + str(line_i), FORMATTING_BLANK(col + str(line_i)))
        sheet[col + str(line_i)].alignment = Alignment(
            textRotation=90,
            horizontal="center"
        )

    for col in COLS_TRANSLATIONS:
        sheet.column_dimensions[col].width = 45
        sheet[col + str(line_i)].border = THIN_BORDER_ALL
        for col in COLS_ATTRIBUTES_ALL[0]:
            sheet[col + str(line_i)].fill = FILL_1B
        for col in COLS_ATTRIBUTES_ALL[1]:
            sheet[col + str(line_i)].fill = FILL_2B
        for col in COLS_ATTRIBUTES_ALL[2]:
            sheet[col + str(line_i)].fill = FILL_3B
        for col in COLS_ATTRIBUTES_ALL[3]:
            sheet[col + str(line_i)].fill = FILL_4B


def add_locked_sheet(workbook, doc_i, doc_k, doc_v):
    sheet = workbook.create_sheet("Orig" + str(doc_i))
    sheet.add_data_validation(VALIDATION_NONE[doc_k + "_orig"])
    sheet.protection.sheet = True

    for col in "ABCDE":
        sheet[col + "1"].font = FONT_BOLD

    # header styling
    # TODO
    sheet["A1"].value = "Source"
    sheet["A1"].fill = FILL_0A
    sheet["B1"].value = "Translation 1"
    sheet["B1"].fill = FILL_1A
    sheet["C1"].value = "Translation 2"
    sheet["C1"].fill = FILL_2A
    sheet["D1"].value = "Translation 3"
    sheet["D1"].fill = FILL_3A
    sheet["E1"].value = "Translation 4"
    sheet["E1"].fill = FILL_4A

    for col in "ABCDE":
        sheet[col + "1"].border = THICK_BORDER_BOTTOM

    sheet.row_dimensions[1].height = 65
    sheet.freeze_panes = sheet["B2"]

    # fill values
    for line_i, line in enumerate(doc_v):
        line_i += 2
        for sent_i, (sent, col) in enumerate(zip(line, "ABCDE")):
            cell = sheet[col + str(line_i)]
            if col == "A" or (line_i - 1 >= DOC_SPANS[doc_k][0] and line_i - 1 <= DOC_SPANS[doc_k][1]):
                cell.value = sent
            cell.alignment = Alignment(wrap_text=True)
            VALIDATION_NONE[doc_k + "_orig"].add(col + str(line_i))
            sheet[col + str(line_i)].font = FONT_NORMAL

        if line_i % 2 == 0:
            sheet["A" + str(line_i)].fill = FILL_0B
            sheet["B" + str(line_i)].fill = FILL_1B
            sheet["C" + str(line_i)].fill = FILL_2B
            sheet["D" + str(line_i)].fill = FILL_3B
            sheet["E" + str(line_i)].fill = FILL_4B

        # set borders & alignments
        for col in "BCDE":
            sheet[col + str(line_i)].border = MEDIUM_BORDER_RIGHT
        for col in COLS_TRANSLATIONS:
            sheet[col + str(line_i)].alignment = Alignment(
                vertical="top", horizontal="justify", wrapText=True,
            )

        sheet["A" + str(line_i)].border = THICK_BORDER_RIGHT
        sheet.row_dimensions[line_i].height = get_height_for_row(sheet, line_i)

    for col in "ABCDE":
        sheet.column_dimensions[col].width = 45


DOC_SPANS = {
    'latimes.431856': (1, 8),
    'metro.co.uk.12069': (6, 13),
    'en.ndtv.com.13152': (3, 10),
    'guardian.260810': (2, 9),
    'rt.com.113909': (1, 8),
    'rt.com.113881': (3, 10),
    'cnn.385674.txt': (1, 8),
    'huffingtonpost.com.19347': (3, 10),
    'brisbanetimes.com.au.225989': (1, 8),
    'express.co.uk.11102': (5, 12),
    'brisbanetimes.com.au.225990': (4, 11),
    'reuters.276709': (4, 11),
    'cbsnews.302129': (3, 10),
    'independent.281139': (1, 8),
    'en.ndtv.com.13143': (3, 10),
    'huffingtonpost.com.19376': (1, 8),
    'cbsnews.302172': (1, 8),
    'huffingtonpost.com.19371.txt': (1, 8),
    'huffingtonpost.com.19385': (1, 8),
    'foxnews.100073': (1, 8),
    'foxnews.100074': (1, 8),
    'scotsman.155294': (1, 8),
}

data = load_data_structure()

# sanitize keys
data = {
    k.strip(): v for k, v in data.items()
}

data = {
    k: v for k, v in data.items()
    if k in DOC_SPANS.keys()
}

for uid_i, uid in enumerate(UIDs[:20]):
    random.seed(uid_i)
    new_data = {}

    keys = list(data.keys())
    random.shuffle(keys)
    print(len(data.keys()), [len(data[k]) for k in keys])

    mapping = {
        "systems": {}, "docs": keys
    }

    for doc_k in keys:
        doc_v = data[doc_k]
        new_data[doc_k] = []

        # shuffle separatedly for each doc
        order = [1, 2, 3, 4]
        random.shuffle(order)
        mapping["systems"][doc_k] = order

        for sent in doc_v:
            new_sent = []
            for i in order:
                new_sent.append(sent[i])

            # source is always first
            new_data[doc_k].append([sent[0]] + new_sent)

    # save mapping
    save_json(f"data/mapping_{uid}.json", mapping)

    # generate XLSX document
    workbook = Workbook()
    # remove default sheet
    workbook.remove(workbook.active)

    for doc_i, (doc_k, doc_v) in enumerate(new_data.items()):
        doc_i += 1
        add_edit_sheet(workbook, doc_i, doc_k, doc_v)
        add_locked_sheet(workbook, doc_i, doc_k, doc_v)

    workbook.save(f"data/translations_{uid}.xlsx")
