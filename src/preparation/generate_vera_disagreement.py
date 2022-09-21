#!/usr/bin/env python3

"""
Creates a XLSX that allows for qualitative analysis by a colaborator
"""

import collections
import itertools
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.text import RichText
import sys
sys.path.append("src")
from xlsx_consts import *
import copy


def transpose(M):
    return [[M[j][i] for j in range(len(M))] for i in range(len(M[0]))]


with open("data/parsed.json", "r") as f:
    data = json.load(f)

print(len(data), type(data))
print(len(data[0]), type(data[0]), data[0].keys())

# categories = list(data[0]["lines"][0]["translations"]["1"]["rating"].keys())
categories = [
    'spelling', 'terminology', 'grammar',
    'meaning', 'style', 'pragmatics', 'overall'
]

workbook_n1_worse = Workbook()
# remove default sheet
workbook_n1_worse.remove(workbook_n1_worse.active)

colors_row_3 = itertools.cycle([FILL_1A, FILL_1B])
colors_row_2 = itertools.cycle([FILL_4A, FILL_4B])
colors_row_0 = itertools.cycle([FILL_0A, FILL_0B])


def next_row_format(rows, first=False):
    row = next(rows)
    color_0 = next(colors_row_0)
    color_2 = next(colors_row_2)
    color_3 = next(colors_row_3)
    for c in row:
        c.alignment = Alignment(
            vertical="top", horizontal="justify", wrapText=True,
        )
    for i in [3, 6]:
        row[i].alignment = Alignment(
            vertical="top", horizontal="center", wrapText=True,
        )

    for i in [1, 2, 3]:
        row[i].fill = color_2
    for i in [4, 5, 6]:
        row[i].fill = color_3
    row[0].fill = color_0

    if first:
        for i in range(7):
            row[i].fill = color_0
            row[i].border = THICK_BORDER_BOTTOM

    return row

SYSTEM_NAME_PRETTY = {
    "1": "P1",
    "2": "P2",
    "3": "P3",
    "4": "N1",
}

data_docs = collections.defaultdict(list)

for category in categories:
    sheet = workbook_n1_worse.create_sheet(category)
    sheet.freeze_panes = sheet["A2"]

    rows = sheet.iter_rows(
        min_row=1, max_row=sys.maxsize,
        min_col=1, max_col=10,
    )

    row = next_row_format(rows, first=True)
    row[0].value = "identifier"
    row[1].value = "Px orig"
    row[2].value = "Px pe"
    row[3].value = "Px score"
    row[4].value = "Px orig"
    row[5].value = "Px pe"
    row[6].value = "Px score"

    for i in range(7):
        row[i].font = FONT_BOLD

    for i in [1, 2, 4, 5]:
        sheet.column_dimensions[ord_to_col(i)].width = 45
    for i in [0, 3, 6]:
        sheet.column_dimensions[ord_to_col(i)].width = 10

    row = next_row_format(rows)

    data_local = []

    for system_name in ["1", "2", "3", "4"]:
        for user_doc in data:
            # we're mutating the object
            user_doc = copy.deepcopy(user_doc)
            for line in user_doc["lines"]:
                line["uid"] = user_doc["uid"]
                line["translation"] = line.pop("translations")[system_name]
            data_docs[
                user_doc["doc"] + "\n" + SYSTEM_NAME_PRETTY[system_name]
            ].append(user_doc["lines"])

        for user_doc_name, user_doc_lines in data_docs.items():
            user_doc_lines = transpose(user_doc_lines)

            for line in user_doc_lines:
                for uid_i in range(len(user_doc_lines)):
                    for uid_j in range(uid_i + 1, len(user_doc_lines)):
                        r1 = line[uid_i]["translation"]["rating"][category]
                        r2 = line[uid_j]["translation"]["rating"][category]
                        if r1 is None or r2 is None:
                            continue
                        if abs(r2 - r1) >= 3:
                            data_local.append((
                                user_doc_name,
                                line[uid_i], line[uid_j],
                                abs(r2 - r1)
                            ))

    print(category, len(data_local))
    # sort by the difference
    data_local.sort(key=lambda x: x[-1], reverse=True)

    for identifier, r1, r2, _ in data_local:
        row[0].value = f"{identifier}\n{r1['uid']}-{r2['uid']}"
        row[1].value = r1["translation"]["orig"]
        row[2].value = r1["translation"]["done"]
        row[3].value = str(r1["translation"]["rating"][category])
        row[4].value = r2["translation"]["orig"]
        row[5].value = r2["translation"]["done"]
        row[6].value = str(r2["translation"]["rating"][category])

        # go to next row
        row = next_row_format(rows)

    for line_i in range(2, sheet.max_row):
        sheet.row_dimensions[line_i].height = get_height_for_row(sheet, line_i)


workbook_n1_worse.save("computed/vera_disagreement.xlsx")
