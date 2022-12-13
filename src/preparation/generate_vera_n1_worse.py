#!/usr/bin/env python3

"""
Creates a XLSX that allows for qualitative analysis by a colaborator
"""

import collections
import itertools
import json
import sys
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.text import RichText
import sys
sys.path.append("src")
from xlsx_consts import *

with open("data/parsed.json", "r") as f:
    data = json.load(f)

print(len(data), type(data))
print(len(data[0]), type(data[0]), data[0].keys())

# categories = list(data[0]["lines"][0]["translations"]["1"]["rating"].keys())
categories = [
    'spelling', 'terminology', 'grammar',
    'meaning', 'style', 'pragmatics', 'overall'
]

workbook_n1_worse = Workbook()
# remove default sheet
workbook_n1_worse.remove(workbook_n1_worse.active)

colors_row_3 = itertools.cycle([FILL_3A, FILL_3B])
colors_row_2 = itertools.cycle([FILL_2A, FILL_2B])
colors_row_0 = itertools.cycle([FILL_0A, FILL_0B])

def next_row_format(rows, first=False):
    row = next(rows)
    color_0 = next(colors_row_0)
    color_2 = next(colors_row_2)
    color_3 = next(colors_row_3)
    for c in row:
        c.alignment = Alignment(
            vertical="top", horizontal="justify", wrapText=True,
        )
    for i in [3, 6]:
        row[i].alignment = Alignment(
            vertical="top", horizontal="center", wrapText=True,
        )

    for i in [1, 2, 3]:
        row[i].fill = color_2
    for i in [4, 5, 6]:
        row[i].fill = color_3
    row[0].fill = color_0

    if first:
        for i in range(7):
            row[i].fill = color_0
            row[i].border = THICK_BORDER_BOTTOM

    return row

n1_worse_all = collections.defaultdict(list)

for category in categories:
    sheet = workbook_n1_worse.create_sheet(category)
    sheet.freeze_panes = sheet["A2"]

    rows = sheet.iter_rows(
        min_row=1, max_row=sys.maxsize,
        min_col=1, max_col=10,
    )

    row = next_row_format(rows, first=True)
    row[0].value = "identifier"
    row[1].value = "N1 orig"
    row[2].value = "N1 pe"
    row[3].value = "N1 score"
    row[4].value = "Px orig"
    row[5].value = "Px pe"
    row[6].value = "Px score"

    for i in range(7):
        row[i].font = FONT_BOLD

    for i in [1, 2, 4, 5]:
        sheet.column_dimensions[ord_to_col(i)].width = 45
    for i in [0, 3, 6]:
        sheet.column_dimensions[ord_to_col(i)].width = 10

    row = next_row_format(rows)

    data_local = []
    for user_doc in data:
        for line in user_doc["lines"]:
            n1 = line["translations"]["4"]
            p1 = line["translations"]["1"]
            p2 = line["translations"]["2"]
            p3 = line["translations"]["3"]

            if p1["rating"]["overall"] is not None and n1["rating"]["overall"] is not None:
                n1_worse_all["p1"].append(p1["rating"]["overall"] > n1["rating"]["overall"])
            if p2["rating"]["overall"] is not None and n1["rating"]["overall"] is not None:
                n1_worse_all["p2"].append(p2["rating"]["overall"] > n1["rating"]["overall"])
            if p3["rating"]["overall"] is not None and n1["rating"]["overall"] is not None:
                n1_worse_all["p3"].append(p3["rating"]["overall"] > n1["rating"]["overall"])

            if any(x["rating"][category] is None for x in [n1, p1, p2, p3]):
                continue

            for px_name, px in zip(["p1", "p2", "p3"], [p1, p2, p3]):
                if n1["rating"][category] >= px["rating"][category]:
                    # not interesting
                    continue
                data_local.append([px_name, n1, px, user_doc, px["rating"][category]-n1["rating"][category]])

    # sort by the difference
    data_local.sort(key=lambda x: x[-1], reverse=True)

    for px_name, n1, px, user_doc, _ in data_local:
        row[0].value = f"{px_name}-{user_doc['uid']}-{user_doc['doc']}"
        row[1].value = n1["orig"]
        row[2].value = n1["done"]
        row[3].value = str(n1["rating"][category])
        row[4].value = px["orig"]
        row[5].value = px["done"]
        row[6].value = str(px["rating"][category])

        # go to next row
        row = next_row_format(rows)

    for line_i in range(2, sheet.max_row):
        sheet.row_dimensions[line_i].height = get_height_for_row(sheet, line_i)

print({k:f"{np.average(v):.2%}" for k,v in n1_worse_all.items()})
workbook_n1_worse.save("computed/vera_n1_worse.xlsx")
